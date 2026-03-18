import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Cities'

headers = ['city', 'state', 'region', 'article_title', 'focus_keyword', 'meta_description', 'article_body']
ws.append(headers)

header_fill = PatternFill('solid', fgColor='E63312')
for cell in ws[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

cities = [
    # HARYANA
    ('Gurgaon','Haryana','NCR','Tile Adhesive in Gurgaon — Best Options {year}','tile adhesive gurgaon','AdhiSeal premium tile adhesive Gurgaon mein available. Strong bonding for vitrified and ceramic tiles.',''),
    ('Faridabad','Haryana','NCR','AdhiSeal Tile Adhesive Dealers in Faridabad {year}','tile adhesive faridabad','Faridabad mein best tile adhesive — AdhiSeal. Contractors aur homeowners ka trusted brand.',''),
    ('Panipat','Haryana','North HR','Tile Adhesive in Panipat | AdhiSeal','tile adhesive panipat','AdhiSeal tile adhesive Panipat ke construction projects ke liye best choice hai.',''),
    ('Sonipat','Haryana','NCR','Best Tile Adhesive in Sonipat for Homes {year}','tile adhesive sonipat','Sonipat ke liye AdhiSeal ke premium tile adhesive solutions. Har tile type ke liye suitable.',''),
    ('Rohtak','Haryana','Central','Tile Adhesive Dealers in Rohtak | AdhiSeal Guide','tile adhesive rohtak','Rohtak mein AdhiSeal ke authorized dealers se premium tile adhesive kharedein.',''),
    ('Hisar','Haryana','West HR','AdhiSeal Tile Adhesive in Hisar — Complete Guide {year}','tile adhesive hisar','Hisar ke liye sahi tile adhesive — AdhiSeal ki product range aur expert guidance.',''),
    ('Karnal','Haryana','North HR','Tile Adhesive in Karnal | Best Brands {year}','tile adhesive karnal','Karnal mein tile adhesive ke liye AdhiSeal — strong grip, easy application.',''),
    ('Ambala','Haryana','North HR','Tile Adhesive Suppliers in Ambala | AdhiSeal','tile adhesive ambala','Ambala ke contractors aur homeowners ke liye AdhiSeal tile adhesive solutions.',''),
    ('Yamunanagar','Haryana','North HR','Best Tile Adhesive in Yamunanagar for {year} Projects','tile adhesive yamunanagar','Yamunanagar mein tile adhesive — AdhiSeal ke premium products ke sath strong bonding.',''),
    ('Rewari','Haryana','South HR','AdhiSeal Tile Adhesive in Rewari | Complete Guide','tile adhesive rewari','Rewari mein AdhiSeal tile adhesive dealers aur product guide.',''),
    ('Bhiwani','Haryana','West HR','Tile Adhesive in Bhiwani — AdhiSeal Products {year}','tile adhesive bhiwani','Bhiwani ke tiling projects ke liye AdhiSeal ke premium adhesive solutions.',''),
    ('Palwal','Haryana','South HR','Best Tile Adhesive Dealers in Palwal | AdhiSeal','tile adhesive palwal','Palwal mein AdhiSeal tile adhesive — quality products, fast delivery.',''),
    ('Jhajjar','Haryana','Central','AdhiSeal Tile Adhesive in Jhajjar | Expert Guide {year}','tile adhesive jhajjar','Jhajjar ke construction projects ke liye sahi tile adhesive chunein — AdhiSeal.',''),
    ('Kurukshetra','Haryana','North HR','Tile Adhesive in Kurukshetra — Complete Buying Guide','tile adhesive kurukshetra','Kurukshetra mein tile adhesive ke liye AdhiSeal — premium quality aur expert support.',''),
    ('Kaithal','Haryana','North HR','AdhiSeal Tile Adhesive Dealers in Kaithal {year}','tile adhesive kaithal','Kaithal mein AdhiSeal tile adhesive ke authorized dealers dhundein.',''),
    ('Jind','Haryana','Central','Tile Adhesive in Jind | AdhiSeal Guide {year}','tile adhesive jind','Jind ke liye best tile adhesive — AdhiSeal ke polymer-modified formulations.',''),
    ('Sirsa','Haryana','West HR','Best Tile Adhesive in Sirsa for Construction Projects','tile adhesive sirsa','Sirsa mein AdhiSeal tile adhesive — strong bonding, easy application.',''),
    ('Fatehabad','Haryana','West HR','AdhiSeal Tile Adhesive Suppliers in Fatehabad | {year}','tile adhesive fatehabad','Fatehabad ke tiling projects ke liye AdhiSeal premium adhesive solutions.',''),
    ('Mahendragarh','Haryana','South HR','Tile Adhesive in Mahendragarh — AdhiSeal Complete Guide','tile adhesive mahendragarh','Mahendragarh mein tile adhesive ke liye AdhiSeal ki trusted product range.',''),
    ('Gurugram','Haryana','NCR','Premium Tile Adhesive in Gurugram for Modern Homes {year}','tile adhesive gurugram','Gurugram ke premium residential aur commercial projects ke liye AdhiSeal Elite adhesive.',''),
    # DELHI
    ('South Delhi','Delhi','Delhi NCR','Tile Adhesive Dealers in South Delhi | AdhiSeal {year}','tile adhesive south delhi','South Delhi ke liye AdhiSeal tile adhesive — premium quality, fast delivery.',''),
    ('North Delhi','Delhi','Delhi NCR','Best Tile Adhesive in North Delhi — AdhiSeal Guide','tile adhesive north delhi','North Delhi mein AdhiSeal ke authorized dealers se tile adhesive kharedein.',''),
    ('East Delhi','Delhi','Delhi NCR','AdhiSeal Tile Adhesive in East Delhi {year}','tile adhesive east delhi','East Delhi ke construction projects ke liye AdhiSeal premium tile adhesive.',''),
    ('West Delhi','Delhi','Delhi NCR','Tile Adhesive Suppliers in West Delhi | Complete Guide','tile adhesive west delhi','West Delhi mein tile adhesive ke liye AdhiSeal — best quality, easy application.',''),
    ('Dwarka','Delhi','Delhi NCR','Tile Adhesive in Dwarka Delhi — Best Options {year}','tile adhesive dwarka delhi','Dwarka ke modern apartments ke liye AdhiSeal Elite tile adhesive — large format tiles.',''),
    ('Rohini','Delhi','Delhi NCR','AdhiSeal Tile Adhesive Dealers in Rohini Delhi','tile adhesive rohini delhi','Rohini ke liye best tile adhesive — AdhiSeal ke premium products aur dealer network.',''),
    ('Pitampura','Delhi','Delhi NCR','Tile Adhesive in Pitampura | AdhiSeal Expert Guide {year}','tile adhesive pitampura','Pitampura ke tiling projects ke liye AdhiSeal tile adhesive — quality guaranteed.',''),
    ('Janakpuri','Delhi','Delhi NCR','Best Tile Adhesive in Janakpuri Delhi {year}','tile adhesive janakpuri','Janakpuri mein AdhiSeal tile adhesive ke authorized dealers dhundein.',''),
    ('Lajpat Nagar','Delhi','Delhi NCR','Tile Adhesive Dealers in Lajpat Nagar Delhi','tile adhesive lajpat nagar','Lajpat Nagar ke liye AdhiSeal tile adhesive — commercial aur residential projects.',''),
    ('Karol Bagh','Delhi','Delhi NCR','AdhiSeal Tile Adhesive in Karol Bagh | {year} Guide','tile adhesive karol bagh','Karol Bagh market ke paas AdhiSeal tile adhesive dealers.',''),
    ('Saket','Delhi','Delhi NCR','Premium Tile Adhesive in Saket Delhi — AdhiSeal','tile adhesive saket delhi','Saket ke premium homes ke liye AdhiSeal Elite tile adhesive.',''),
    ('Vasant Kunj','Delhi','Delhi NCR','Tile Adhesive in Vasant Kunj — AdhiSeal Complete Guide','tile adhesive vasant kunj','Vasant Kunj ke premium residential projects ke liye AdhiSeal tile adhesive.',''),
    ('Noida','Delhi','Delhi NCR','Best Tile Adhesive in Noida Sector | AdhiSeal {year}','tile adhesive noida','Noida ke modern apartments ke liye AdhiSeal tile adhesive.',''),
    ('Ghaziabad','Delhi','Delhi NCR','Tile Adhesive Dealers in Ghaziabad | AdhiSeal Guide','tile adhesive ghaziabad','Ghaziabad ke construction projects ke liye AdhiSeal premium tile adhesive.',''),
    # UTTAR PRADESH
    ('Agra','Uttar Pradesh','UP','Tile Adhesive in Agra | AdhiSeal Products {year}','tile adhesive agra','Agra ke construction projects ke liye AdhiSeal tile adhesive.',''),
    ('Lucknow','Uttar Pradesh','UP','AdhiSeal Tile Adhesive Dealers in Lucknow {year}','tile adhesive lucknow','Lucknow mein tile adhesive ke liye AdhiSeal — premium quality.',''),
    ('Kanpur','Uttar Pradesh','UP','Best Tile Adhesive in Kanpur — AdhiSeal Guide {year}','tile adhesive kanpur','Kanpur ke liye AdhiSeal tile adhesive — contractors aur homeowners ka choice.',''),
    ('Meerut','Uttar Pradesh','NCR','Tile Adhesive in Meerut | AdhiSeal Expert Guide','tile adhesive meerut','Meerut aur NCR belt ke liye AdhiSeal tile adhesive solutions.',''),
    # PUNJAB
    ('Ludhiana','Punjab','Punjab','Best Tile Adhesive in Ludhiana | AdhiSeal {year}','tile adhesive ludhiana','Ludhiana ke liye AdhiSeal tile adhesive — industrial aur residential.',''),
    ('Amritsar','Punjab','Punjab','AdhiSeal Tile Adhesive Dealers in Amritsar | Guide','tile adhesive amritsar','Amritsar mein tile adhesive ke liye AdhiSeal ke premium products.',''),
    ('Jalandhar','Punjab','Punjab','Tile Adhesive in Jalandhar — AdhiSeal Complete Guide {year}','tile adhesive jalandhar','Jalandhar ke construction projects ke liye AdhiSeal tile adhesive.',''),
    # RAJASTHAN
    ('Jaipur','Rajasthan','Rajasthan','Tile Adhesive in Jaipur | AdhiSeal Products {year}','tile adhesive jaipur','Jaipur ke construction projects ke liye AdhiSeal.',''),
    ('Jodhpur','Rajasthan','Rajasthan','AdhiSeal Tile Adhesive Dealers in Jodhpur | Guide','tile adhesive jodhpur','Jodhpur ke buildings ke liye AdhiSeal tile adhesive.',''),
    ('Udaipur','Rajasthan','Rajasthan','Best Tile Adhesive in Udaipur | AdhiSeal Guide','tile adhesive udaipur','Udaipur ke premium hotels aur homes ke liye AdhiSeal Elite adhesive.',''),
]

for c in cities:
    ws.append(list(c))

ws.column_dimensions['A'].width = 18
ws.column_dimensions['B'].width = 16
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 52
ws.column_dimensions['E'].width = 28
ws.column_dimensions['F'].width = 60
ws.column_dimensions['G'].width = 60

out = 'data/north_india_cities.xlsx'
wb.save(out)
print(f'Done! {out} created with {len(cities)} cities.')
