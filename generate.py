#!/usr/bin/env python3
"""
=======================================================
 AdhiSeal Blog — Automated Generator (v2)
 blog.adhiseal.com

 Two-Layer Architecture:
   Layer 1 — Blog Posts (visible on homepage/nav)
             Source: data/posts.json
             Output: /output/posts/{slug}.html
                     /output/category/{cat}.html
                     /output/index.html

   Layer 2 — Geo Pages (sitemap only, NOT on homepage)
             Source: data/north_india_cities.xlsx
             Output: /output/geo/{state}/{city}.html

 Usage:
   python generate.py                    # full build
   python generate.py --only-blog        # blog posts only
   python generate.py --only-geo         # geo pages only
   python generate.py --no-submit        # skip sitemap ping
   python generate.py --clean            # clean output first

=======================================================
"""

import os, re, sys, json, argparse, shutil, logging
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("❌  openpyxl not found. Run: pip install openpyxl jinja2 requests")
    sys.exit(1)

try:
    from jinja2 import Environment, FileSystemLoader
except ImportError:
    print("❌  jinja2 not found. Run: pip install openpyxl jinja2 requests")
    sys.exit(1)

try:
    import requests
except ImportError:
    requests = None

# ---- Config ----
BASE_DIR      = Path(__file__).parent
TEMPLATES_DIR = BASE_DIR / "templates"
OUTPUT_DIR    = BASE_DIR / "output"
ASSETS_SRC    = BASE_DIR / "assets"
DATA_DIR      = BASE_DIR / "data"
BLOG_DOMAIN   = "https://blog.adhiseal.com"
YEAR          = datetime.now().year
TODAY         = datetime.now().strftime("%Y-%m-%d")
TODAY_DISPLAY = datetime.now().strftime("%d %B %Y")

logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger(__name__)

MULTI_CATEGORIES = {
    "manufacturer": "Manufacturer",
    "supplier": "Supplier",
    "job-work": "Job Work"
}

MULTI_PRODUCTS = {
    "tile-adhesive": "Tile Adhesive",
    "pu-adhesive": "PU Adhesive",
    "epoxy-grout": "Epoxy Grout",
    "tile-grout": "Tile Grout",
    "grout-admix": "Grout Admix",
    "grout-hardener": "Grout Hardener",
    "tile-leveler": "Tile Leveler",
    "spacer": "Spacer",
    "aac-block-adhesive": "AAC Block Adhesive",
    "waterproofing-chemical": "Waterproofing Chemical",
    "tile-cleaner": "Tile Cleaner",
    "sbr-latex": "SBR Latex",
    "araldite": "Araldite",
    "epoxy-resin": "Epoxy Resin",
    "epoxy-hardner": "Epoxy Hardner"
}

MULTI_TEMPLATES = {
    "manufacturer": """<p style="text-align: center;"><img src="{css_path}assets/products/{product_slug}.jpg" alt="{product_name} in {city}" style="max-width: 100%; border-radius: 8px; box-shadow: 0 4px 6px rgba(0,0,0,0.1);" onerror="this.style.display='none'"></p>

Are you looking for a reliable and top-quality **{product_name} Manufacturer** in {city}? Whether it's for a commercial project or a residential buildup, {brand} offers the best-in-class {product_name} right here in {state}.

## Why Choose {brand} as your {product_name} Manufacturer in {city}?

Finding the right manufacturer for {product_name} in {city} ({pincode}) is critical for the durability and strength of your construction projects. As a leading manufacturer in {state}, {brand} ensures:
- **Premium Quality:** Formulated with advanced polymers and raw materials.
- **Trusted by Experts:** Widely used by contractors and architects across {city}.
- **Consistent Supply:** Timely delivery directly from our manufacturing units to your site.

### Our Manufacturing Standards

Every batch of {product_name} manufactured by {brand} undergoes rigorous testing. We understand the local climate and building requirements of {city}, ensuring our products perform exceptionally well under local conditions.

## Contact the Best {product_name} Manufacturer Today!

Don't compromise on quality. Reach out to {brand} for bulk orders and distributor inquiries in {city}. Get the best pricing directly from the manufacturer!
""",
    "supplier": """<p style="text-align: center;"><img src="{css_path}assets/products/{product_slug}.jpg" alt="{product_name} in {city}" style="max-width: 100%; border-radius: 8px; box-shadow: 0 4px 6px rgba(0,0,0,0.1);" onerror="this.style.display='none'"></p>

If you need a trusted **{product_name} Supplier** in {city}, you are at the right place. {brand} provides seamless supply of premium {product_name} across {city}, {state} ({pincode}).

## Top {product_name} Supplier in {city}

As a leading supplier of building materials and chemicals in {state}, {brand} provides wholesale and retail supply of {product_name} for all your tiling and construction needs.

### Why Source {product_name} from us?
- **Ready Stock:** We maintain high inventory levels to meet immediate demands in {city}.
- **Competitive Pricing:** Get the best supplier rates directly from {brand}.
- **Authentic Products:** 100% genuine {brand} {product_name} guaranteed.
- **Fast Delivery:** Our strong local network ensures rapid delivery to your project site in {pincode}.

## Build Stronger with {brand}

Contractors, builders, and home-owners in {city} trust us as their go-to supplier for {product_name}. Partner with {brand} today for uninterrupted supply!
""",
    "job-work": """<p style="text-align: center;"><img src="{css_path}assets/products/{product_slug}.jpg" alt="{product_name} in {city}" style="max-width: 100%; border-radius: 8px; box-shadow: 0 4px 6px rgba(0,0,0,0.1);" onerror="this.style.display='none'"></p>

Looking for professional **{product_name} Job Work** services in {city}? {brand} provides expert application and job work solutions for {product_name} across {state}.

## Professional {product_name} Job Work in {city}

Having the right product is only half the job; applying it correctly is what ensures long-lasting results. {brand} offers specialized job work and application services for {product_name} in {city} ({pincode}).

### Why Contract Our Job Work Services?
- **Experienced Applicators:** Our team consists of trained professionals who know exactly how to handle {product_name}.
- **Flawless Finish:** We guarantee a high-quality, durable finish for every project.
- **Time-Saving:** Efficient execution means your project stays on schedule.
- **End-to-End Solution:** From supplying the premium {brand} {product_name} to executing the job work in {city}.

## Request a Job Work Quote Today!

Whether it's a large commercial site or a residential project in {city}, {brand} has the expertise to handle your {product_name} application flawlessly. Get in touch with us for a free site assessment and quote!
"""
}

# ===================================================================
# UTILITIES
# ===================================================================

def slugify(text: str) -> str:
    text = str(text).lower().strip()
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'[\s_]+', '-', text)
    return text.strip('-')

def read_time(text: str) -> int:
    return max(2, round(len(re.findall(r'\w+', text)) / 200))

def wrap_md(text: str) -> str:
    """Convert simple markdown to HTML (headings, bold, lists, tables)."""
    lines = text.split('\n')
    html = []
    in_list = False
    in_table = False

    for line in lines:
        stripped = line.strip()

        # Table row
        if stripped.startswith('|') and stripped.endswith('|'):
            if not in_table:
                html.append('<table>')
                in_table = True
                cells = [c.strip() for c in stripped[1:-1].split('|')]
                html.append('<tr>' + ''.join(f'<th>{c}</th>' for c in cells) + '</tr>')
                continue
            # separator row
            if all(c.strip().replace('-','').replace(':','') == '' for c in stripped[1:-1].split('|')):
                continue
            cells = [c.strip() for c in stripped[1:-1].split('|')]
            html.append('<tr>' + ''.join(f'<td>{c}</td>' for c in cells) + '</tr>')
            continue
        else:
            if in_table:
                html.append('</table>')
                in_table = False

        # Close list
        if in_list and not stripped.startswith('- ') and not stripped.startswith('* '):
            html.append('</ul>')
            in_list = False

        # Headings
        m2 = re.match(r'^## (.+)', line)
        m3 = re.match(r'^### (.+)', line)
        if m2:
            htext = m2.group(1).strip()
            hid = slugify(htext)
            html.append(f'<h2 id="{hid}">{htext}</h2>')
        elif m3:
            htext = m3.group(1).strip()
            hid = slugify(htext)
            html.append(f'<h3 id="{hid}">{htext}</h3>')
        elif stripped.startswith('- ') or stripped.startswith('* '):
            if not in_list:
                html.append('<ul>')
                in_list = True
            content = stripped[2:]
            content = re.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', content)
            html.append(f'<li>{content}</li>')
        elif stripped == '':
            html.append('')
        else:
            line_out = re.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', stripped)
            line_out = re.sub(r'\*(.+?)\*', r'<em>\1</em>', line_out)
            html.append(f'<p>{line_out}</p>')

    if in_list:
        html.append('</ul>')
    if in_table:
        html.append('</table>')

    return '\n'.join(html)

def build_toc(html: str) -> str:
    items = []
    for m in re.finditer(r'<(h[23]) id="([^"]+)">([^<]+)</', html):
        tag, hid, text = m.group(1), m.group(2), m.group(3)
        cls = ' class="h3"' if tag == 'h3' else ''
        items.append(f'<li{cls}><a href="#{hid}">{text}</a></li>')
    return '\n'.join(items)

def expand(text: str, city='', state='', region='') -> str:
    for k, v in {'{city}': city, '{state}': state, '{region}': region or state,
                 '{year}': str(YEAR), '{brand}': 'AdhiSeal', '{date}': TODAY_DISPLAY}.items():
        text = text.replace(k, v)
    return text


# ===================================================================
# LAYER 1 — BLOG POSTS
# ===================================================================

def read_posts() -> list[dict]:
    path = DATA_DIR / "posts.json"
    if not path.exists():
        log.warning("⚠️   posts.json not found — no blog posts will be generated")
        return []
    with open(path, encoding='utf-8') as f:
        posts = json.load(f)
    log.info(f"📝  Loaded {len(posts)} blog posts from posts.json")
    return posts

def render_post_card(post: dict, css_path: str = '') -> str:
    slug = post['slug']
    title = post['title']
    excerpt = post.get('excerpt', '')[:120]
    cat = post.get('category', '')
    cat_label = post.get('category_label', cat)
    date_display = post.get('date_display', TODAY_DISPLAY)
    rt = post.get('read_time', 3)
    color = post.get('image_color', '#E63312')
    initials = title[:3].upper().replace(' ', '')
    url = f"{css_path}posts/{slug}/"

    return f'''<div class="post-card animate-up" data-cat="{cat}" data-title="{title.lower()}" data-excerpt="{excerpt.lower()}" onclick="location.href='{url}'">
  <div class="post-card-img">
    <div class="post-card-img-bg" style="--img-color:{color};"></div>
    <div class="post-card-img-text">{initials}</div>
    <span class="post-card-badge">{cat_label}</span>
  </div>
  <div class="post-card-body">
    <div class="post-card-cat">{cat_label}</div>
    <h3 class="post-card-title">{title}</h3>
    <p class="post-card-excerpt">{excerpt}...</p>
  </div>
  <div class="post-card-footer">
    <span>📅 {date_display}</span>
    <span>⏱ {rt} min</span>
  </div>
</div>'''

def render_post_page(post: dict, all_posts: list[dict], jinja_env) -> str:
    body_html = wrap_md(post.get('body', ''))
    toc_html = build_toc(body_html)
    related = [p for p in all_posts if p['category'] == post['category'] and p['slug'] != post['slug']][:3]

    template = jinja_env.get_template('post.html')
    return template.render(
        post_title         = post['title'],
        post_slug          = post['slug'],
        post_excerpt       = post.get('excerpt', ''),
        post_category      = post.get('category', ''),
        post_category_label= post.get('category_label', ''),
        post_date          = post.get('date', TODAY),
        post_date_display  = post.get('date_display', TODAY_DISPLAY),
        post_author        = post.get('author', 'AdhiSeal Team'),
        post_read_time     = post.get('read_time', read_time(post.get('body', ''))),
        post_image_color   = post.get('image_color', '#E63312'),
        post_tags          = post.get('tags', []),
        tags_str           = ', '.join(post.get('tags', [])),
        post_body_html     = body_html,
        toc_html           = toc_html,
        related_posts      = related,
        year               = YEAR,
    )

def render_category_page(cat: str, posts: list[dict], all_posts: list[dict], jinja_env) -> str:
    """Render a category listing page using index.html template."""
    cat_labels = {p['category']: p.get('category_label', p['category']) for p in all_posts}
    cat_label = cat_labels.get(cat, cat)
    cards = ''.join(render_post_card(p, css_path='../../') for p in posts)
    featured = next((p for p in posts if p.get('featured')), posts[0] if posts else None)
    popular_html = ''
    for i, p in enumerate(all_posts[:5], 1):
        popular_html += f'<li><span class="widget-num">{i}</span><a href="../../posts/{p["slug"]}/">{p["title"]}</a></li>'

    counts = {}
    for p in all_posts:
        counts[p['category']] = counts.get(p['category'], 0) + 1

    template = jinja_env.get_template('index.html')
    return template.render(
        page_title        = f"{cat_label} — AdhiSeal Blog",
        page_description  = f"AdhiSeal\'s {cat_label} articles — tile adhesive tips, guides, and expert advice.",
        canonical         = f"category/{cat}/",
        hero_title        = f"<span>{cat_label}</span>",
        hero_subtitle     = f"AdhiSeal\'s {cat_label} — practical knowledge about tile adhesives.",
        total_posts       = len(all_posts),
        total_cities      = 44,
        featured_post     = featured,
        post_cards_html   = cards,
        popular_posts_html= popular_html,
        cat_counts        = {'tips': counts.get('tips', 0), 'guides': counts.get('guides', 0), 'products': counts.get('products', 0)},
        css_path          = '../../',
        year              = YEAR,
    )

def build_blog_layer(all_posts: list[dict], jinja_env) -> list[str]:
    """Generate all blog post pages, category pages, and main index. Returns URLs."""
    if not all_posts:
        return []

    posts_dir = OUTPUT_DIR / 'posts'
    posts_dir.mkdir(exist_ok=True)
    cat_dir = OUTPUT_DIR / 'category'
    cat_dir.mkdir(exist_ok=True)

    urls = [f"{BLOG_DOMAIN}/"]

    # Post pages
    for post in all_posts:
        html = render_post_page(post, all_posts, jinja_env)
        post_dir = posts_dir / post['slug']
        post_dir.mkdir(exist_ok=True)
        (post_dir / "index.html").write_text(html, encoding='utf-8')
        urls.append(f"{BLOG_DOMAIN}/posts/{post['slug']}/")
        log.info(f"  📄  posts/{post['slug']}/index.html")

    # Category pages
    categories = {}
    for p in all_posts:
        cat = p.get('category', 'misc')
        categories.setdefault(cat, []).append(p)

    for cat, posts in categories.items():
        html = render_category_page(cat, posts, all_posts, jinja_env)
        cat_folder = cat_dir / cat
        cat_folder.mkdir(exist_ok=True)
        (cat_folder / "index.html").write_text(html, encoding='utf-8')
        urls.append(f"{BLOG_DOMAIN}/category/{cat}/")
        log.info(f"  📂  category/{cat}/index.html ({len(posts)} posts)")

    # Main index
    featured = next((p for p in all_posts if p.get('featured')), all_posts[0])
    cards = ''.join(render_post_card(p) for p in all_posts if not p.get('featured'))
    popular_html = ''
    for i, p in enumerate(all_posts[:5], 1):
        popular_html += f'<li><span class="widget-num">{i}</span><a href="posts/{p["slug"]}/">{p["title"]}</a></li>'
    counts = {}
    for p in all_posts:
        counts[p['category']] = counts.get(p['category'], 0) + 1

    template = jinja_env.get_template('index.html')
    index_html = template.render(
        page_title        = "AdhiSeal Blog — Tile Adhesive Tips & Expert Guides",
        page_description  = "Expert guides, comparisons, tips, and installation advice about tile adhesives from the AdhiSeal blog.",
        canonical         = "",
        hero_title        = "Tile Adhesive Tips &amp; <span>Expert Guides</span>",
        hero_subtitle     = "For contractors, homeowners, and architects — from tile selection to installation. Learn from the AdhiSeal experts.",
        total_posts       = len(all_posts),
        total_cities      = 44,
        featured_post     = featured,
        post_cards_html   = cards,
        popular_posts_html= popular_html,
        cat_counts        = {'tips': counts.get('tips', 0), 'guides': counts.get('guides', 0), 'products': counts.get('products', 0)},
        css_path          = '',
        year              = YEAR,
    )
    (OUTPUT_DIR / 'index.html').write_text(index_html, encoding='utf-8')
    log.info("🏠  index.html generated")

    return urls


# ===================================================================
# LAYER 2 — GEO PAGES
# ===================================================================

DEFAULT_GEO_ARTICLE = """Choosing the right tile adhesive in {city} is an important decision — whether it's a home renovation or a new construction. {brand} products are widely available in {city}, {state} and are trusted by thousands of contractors.

## Why Quality Tile Adhesive Matters in {city}

The quality of tile adhesive matters a lot in {city}'s construction projects. Depending on the climate and building conditions in {state}, selecting the right adhesive is zaroori (essential).

### Strong Bonding

Homes and commercial spaces in {city} require a strong, water-resistant adhesive for tiles. {brand}'s polymer-modified adhesives are perfect for this job.

### Wide Range

Floor tiles, wall tiles, large format — {brand} has the right grade available for every application.

## {brand} Product Range — Best Options for {city}

### {brand} Premium Adhesive
Ideal for interior walls and floors. Provides the best results with ceramic and vitrified tiles. A trusted choice for residential projects in {city}.

### {brand} Standard+ Adhesive
A smooth-applying formula for everyday ceramic and vitrified tile installations. Perfect for medium-budget projects in {state}.

### {brand} Super Adhesive
High-flexibility formula — designed for tough conditions and high-movement areas. Widely used in {city}'s commercial projects.

### {brand} Elite Adhesive
Heavy-duty application for large format tiles (above 600x600mm). For premium residential and commercial projects in {state}.

## Where to Find {brand} in {city}?

{brand}'s authorized dealers are available across {city} and {state}. To find your nearest dealer, visit: adhiseal.com/dealer/

## Tile Installation Tips for {city}

- **Surface preparation**: The surface should be clean and dry.
- **Coverage**: One bag covers about 4-6 sq.meters.
- **Open time**: Use within 30-45 minutes after mixing.
- **Curing**: Grout after 24 hours, achieves full strength in 48-72 hours.

## Why Choose {brand}?

- Premium polymer-modified formula
- Trusted brand for {state} contractors
- Wide product range for every tile type
- Technical support available
- Fast delivery through our dealer network in {city}

Get in touch with {brand} for your {city} projects and give your tiling a professional finish.
"""

def read_cities_file(filepath: Path) -> list[dict]:
    if filepath.suffix.lower() == '.csv':
        import csv
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            headers = [str(h).strip().lower() for h in next(reader)]
            rows = []
            for row in reader:
                if not any(row): continue
                data = {headers[j]: (str(v).strip() if v is not None else '') for j, v in enumerate(row) if j < len(headers)}
                if data.get('city'):
                    rows.append(data)
            log.info(f"🗺️   Loaded {len(rows)} cities from CSV")
            return rows
    else:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else '' for c in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            data = {headers[j]: (str(v).strip() if v is not None else '') for j, v in enumerate(row)}
            if data.get('city'):
                rows.append(data)
        log.info(f"🗺️   Loaded {len(rows)} cities from Excel")
        return rows

def render_geo_page(row: dict, all_rows: list[dict], jinja_env, category="Manufacturing", product="Tile-Adhesive") -> tuple[str, str, str]:
    city   = row.get('city', '').strip()
    state  = row.get('state', '').strip()
    region = row.get('region', '').strip()
    pincode = row.get('pincode', '').strip()

    if product == 'Tile-Manufacturers':
        default_title = f'Tile Manufacturers in {city}, {region}-{pincode}'
        default_meta = f'Find the best tile manufacturers in {city}, {region}. Top quality tile manufacturing in {pincode}.'
        default_kw = f'tile manufacturers {city}, tile manufacturers {region} {pincode}'
    else:
        default_title = f'Tile Adhesive in {city} | AdhiSeal'
        default_meta = f'AdhiSeal tile adhesive {city} mein available. Strong bonding.'
        default_kw = f'tile adhesive {city}'

    title   = expand(row.get('article_title', default_title), city, state, region)
    meta    = expand(row.get('meta_description', default_meta), city, state, region)
    kw      = expand(row.get('focus_keyword', default_kw), city, state, region)
    raw     = expand(row.get('article_body', '') or DEFAULT_GEO_ARTICLE, city, state, region)

    city_slug  = slugify(city)
    state_slug = slugify(state)

    body_html = wrap_md(raw)
    toc_html  = build_toc(body_html)
    rt        = read_time(raw)

    # Related — same state
    related = [r for r in all_rows if slugify(r.get('state','')) == state_slug and slugify(r.get('city','')) != city_slug][:3]
    related_html = ''
    for r in related:
        rc_slug = slugify(r.get('city',''))
        related_html += f'<div class="related-card" onclick="location.href=\'../{rc_slug}/\'"><div class="related-card-cat">{state}</div><div class="related-card-title">{r.get("city","")} — Tile Adhesive Guide</div></div>'

    # State sidebar
    state_cities = [r for r in all_rows if slugify(r.get('state','')) == state_slug][:12]
    state_cities_html = ''
    for r in state_cities:
        rc_slug = slugify(r.get('city',''))
        active = 'style="color:var(--primary);font-weight:700;"' if rc_slug == city_slug else ''
        state_cities_html += f'<li><span class="widget-num">📍</span><a href="../{rc_slug}/" {active}>{r.get("city","")}</a></li>'

    template = jinja_env.get_template('article.html')
    html = template.render(
        article_title      = title,
        meta_description   = meta,
        focus_keyword      = kw,
        city               = city,
        state              = state,
        city_slug          = city_slug,
        state_slug         = state_slug,
        category           = category,
        product            = product,
        date_published     = TODAY_DISPLAY,
        article_body_html  = body_html,
        toc_html           = toc_html,
        read_time          = rt,
        related_posts_html = related_html,
        state_cities_html  = state_cities_html,
        css_path           = "../../../../",
        canonical          = f"{category}/{product}/{state_slug}/{city_slug}/",
        show_hero_image    = True,
        year               = YEAR,
    )
    return state_slug, city_slug, html

def build_geo_layer(all_rows: list[dict], jinja_env, category="Manufacturing", product="Tile-Adhesive") -> list[str]:
    """Generate geo pages. Returns URLs for sitemap."""
    if not all_rows:
        return []

    geo_dir = OUTPUT_DIR / category / product
    geo_dir.mkdir(parents=True, exist_ok=True)
    urls = []

    state_rows: dict[str, list] = {}
    for row in all_rows:
        state_slug, city_slug, html = render_geo_page(row, all_rows, jinja_env, category, product)
        state_dir = geo_dir / state_slug
        city_dir = state_dir / city_slug
        city_dir.mkdir(parents=True, exist_ok=True)
        (city_dir / "index.html").write_text(html, encoding='utf-8')
        urls.append(f"{BLOG_DOMAIN}/{category}/{product}/{state_slug}/{city_slug}/")
        state_rows.setdefault(state_slug, []).append(row)

    log.info(f"🗺️   Generated {len(urls)} geo pages under output/{category}/{product}/")
    return urls


# ===================================================================
# LAYER 4 — MULTI PRODUCT & CATEGORY PAGES
# ===================================================================

def render_multi_page(row: dict, all_rows: list[dict], jinja_env, cat_slug: str, cat_name: str, prod_slug: str, prod_name: str) -> tuple[str, str]:
    city   = row.get('city', '').strip()
    state  = row.get('state', '').strip()
    pincode = row.get('pincode', '').strip()

    title   = f"{prod_name} {cat_name} in {city} | AdhiSeal"
    meta    = f"Find the best {prod_name} {cat_name} in {city}, {state} {pincode}. High-quality products and services by AdhiSeal."
    kw      = f"{prod_name} {cat_name} {city}"
    
    custom_template_path = DATA_DIR / "multi_templates" / f"{cat_slug}_{prod_slug}.md"
    if custom_template_path.exists():
        raw = custom_template_path.read_text(encoding='utf-8')
    else:
        raw = MULTI_TEMPLATES.get(cat_slug, "")
    for k, v in {'{city}': city, '{state}': state, '{pincode}': pincode, '{product_name}': prod_name, '{product_slug}': prod_slug, '{brand}': 'AdhiSeal', '{css_path}': '../../../'}.items():
        raw = raw.replace(k, v)

    city_slug  = slugify(city)

    body_html = wrap_md(raw)
    toc_html  = build_toc(body_html)
    rt        = read_time(raw)

    # State sidebar (related cities)
    state_cities = [r for r in all_rows if slugify(r.get('state','')) == slugify(state)][:12]
    state_cities_html = ''
    for r in state_cities:
        rc_slug = slugify(r.get('city',''))
        active = 'style="color:var(--primary);font-weight:700;"' if rc_slug == city_slug else ''
        state_cities_html += f'<li><span class="widget-num">📍</span><a href="../{rc_slug}/" {active}>{r.get("city","")}</a></li>'

    template = jinja_env.get_template('article.html')
    html = template.render(
        article_title      = title,
        meta_description   = meta,
        focus_keyword      = kw,
        city               = city,
        state              = state,
        city_slug          = city_slug,
        state_slug         = slugify(state),  # we pass it though not in url
        category           = cat_name,
        product            = prod_name,
        date_published     = TODAY_DISPLAY,
        article_body_html  = body_html,
        toc_html           = toc_html,
        read_time          = rt,
        related_posts_html = "",
        state_cities_html  = state_cities_html,
        css_path           = "../../../",
        canonical          = f"{cat_slug}/{prod_slug}/{city_slug}/",
        show_hero_image    = False,
        year               = YEAR,
    )
    return city_slug, html

def build_multi_layer(all_rows: list[dict], jinja_env) -> list[str]:
    if not all_rows:
        return []

    urls = []
    total = len(all_rows) * len(MULTI_CATEGORIES) * len(MULTI_PRODUCTS)

    for cat_slug, cat_name in MULTI_CATEGORIES.items():
        for prod_slug, prod_name in MULTI_PRODUCTS.items():
            dir_path = OUTPUT_DIR / cat_slug / prod_slug
            dir_path.mkdir(parents=True, exist_ok=True)
            for row in all_rows:
                city_slug, html = render_multi_page(row, all_rows, jinja_env, cat_slug, cat_name, prod_slug, prod_name)
                city_dir = dir_path / city_slug
                city_dir.mkdir(parents=True, exist_ok=True)
                (city_dir / "index.html").write_text(html, encoding='utf-8')
                urls.append(f"{BLOG_DOMAIN}/{cat_slug}/{prod_slug}/{city_slug}/")

    log.info(f"🚀  Generated {len(urls)} multi-product pages.")
    return urls


# ===================================================================
# SITEMAP
# ===================================================================

def generate_sitemap(blog_urls: list[str], geo_urls: list[str]) -> None:
    all_urls = blog_urls + geo_urls
    items = []
    for url in all_urls:
        is_index = url.endswith('index.html') and url.count('/') <= 4
        if not is_index and url == f"{BLOG_DOMAIN}/":
            is_index = True
        is_post  = '/posts/' in url
        priority = '1.0' if is_index else ('0.8' if is_post else '0.5')
        freq     = 'daily' if is_index or is_post else 'weekly'
        items.append(
            f'  <url>\n    <loc>{url}</loc>\n    <lastmod>{TODAY}</lastmod>\n'
            f'    <changefreq>{freq}</changefreq>\n    <priority>{priority}</priority>\n  </url>'
        )

    sitemap = f'''<?xml version="1.0" encoding="UTF-8"?>
<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">
  <!-- AdhiSeal Blog — Generated {TODAY} | {len(blog_urls)} blog + {len(geo_urls)} geo = {len(all_urls)} total URLs -->
{chr(10).join(items)}
</urlset>'''

    (OUTPUT_DIR / 'sitemap.xml').write_text(sitemap, encoding='utf-8')
    log.info(f"🗺️   sitemap.xml — {len(all_urls)} URLs ({len(blog_urls)} blog + {len(geo_urls)} geo)")


# ===================================================================
# SITEMAP SUBMISSION
# ===================================================================

def submit_sitemap():
    if requests is None:
        log.warning("⚠️   requests not installed — skipping")
        return
    sitemap_url = f"{BLOG_DOMAIN}/sitemap.xml"
    engines = {
        "Google": f"https://www.google.com/ping?sitemap={sitemap_url}",
        "Bing":   f"https://www.bing.com/ping?sitemap={sitemap_url}",
    }
    log.info("🚀  Submitting sitemap...")
    for name, url in engines.items():
        try:
            r = requests.get(url, timeout=10)
            log.info(f"  {'✅' if r.status_code in (200,202) else '⚠️'} {name}: HTTP {r.status_code}")
        except Exception as e:
            log.warning(f"  ❌  {name}: {e}")


# ===================================================================
# MAIN
# ===================================================================

def main():
    parser = argparse.ArgumentParser(description="AdhiSeal Blog Generator v2")
    parser.add_argument('--input',      default='data/north_india_cities.xlsx')
    parser.add_argument('--mfg-csv',    default=None, help='Generate Tile Manufacturers pages from this CSV')
    parser.add_argument('--multi',      action='store_true', help='Generate programmatic Multi-Product pages')
    parser.add_argument('--no-submit',  action='store_true')
    parser.add_argument('--clean',      action='store_true')
    parser.add_argument('--only-blog',  action='store_true', help='Generate blog layer only')
    parser.add_argument('--only-geo',   action='store_true', help='Generate geo layer only')
    args = parser.parse_args()

    log.info("=" * 55)
    log.info("  AdhiSeal Blog Generator v2 — Starting...")
    log.info("=" * 55)

    if args.clean and OUTPUT_DIR.exists():
        shutil.rmtree(OUTPUT_DIR)
        log.info("🧹  Cleaned output/")

    OUTPUT_DIR.mkdir(exist_ok=True)

    if ASSETS_SRC.exists():
        shutil.copytree(ASSETS_SRC, OUTPUT_DIR / 'assets', dirs_exist_ok=True)
        log.info("📁  Assets copied")

    jinja_env = Environment(loader=FileSystemLoader(str(TEMPLATES_DIR)))

    blog_urls = []
    geo_urls  = []

    # --- Layer 1: Blog Posts ---
    if not args.only_geo:
        log.info("\n📝  Building BLOG layer...")
        posts = read_posts()
        if posts:
            blog_urls = build_blog_layer(posts, jinja_env)
            log.info(f"  ✅  Blog: {len(posts)} posts + categories + homepage")
        else:
            log.warning("  ⚠️   No posts found — homepage will be empty")

    # --- Layer 2: Geo Pages ---
    if not args.only_blog:
        input_path = BASE_DIR / args.input
        if input_path.exists():
            log.info("\n🗺️   Building GEO layer...")
            cities = read_cities_file(input_path)
            geo_urls = build_geo_layer(cities, jinja_env)
            log.info(f"  ✅  Geo: {len(geo_urls)} city pages (hidden from homepage)")
        else:
            log.warning(f"  ⚠️   {input_path} not found — geo pages skipped")

    # --- Layer 3: Manufacturers ---
    if getattr(args, 'mfg_csv', None):
        mfg_path = Path(args.mfg_csv)
        if not mfg_path.is_absolute():
            mfg_path = BASE_DIR / args.mfg_csv
        if mfg_path.exists():
            log.info("\n🏭   Building MANUFACTURERS layer from CSV...")
            mfg_cities = read_cities_file(mfg_path)
            mfg_urls = build_geo_layer(mfg_cities, jinja_env, category="Manufacturing", product="Tile-Manufacturers")
            geo_urls.extend(mfg_urls)
            log.info(f"  ✅  Mfg: {len(mfg_urls)} manufacturer city pages via CSV")
        else:
            log.warning(f"  ⚠️   {mfg_path} not found — mfg pages skipped")

    # --- Layer 4: Multi-Product Pages ---
    if args.multi:
        input_path = BASE_DIR / args.input
        if input_path.exists():
            log.info("\n🚀   Building MULTI-PRODUCT layer...")
            cities = read_cities_file(input_path)
            multi_urls = build_multi_layer(cities, jinja_env)
            geo_urls.extend(multi_urls)
            log.info(f"  ✅  Multi: {len(multi_urls)} programmatic pages generated")
        else:
            log.warning(f"  ⚠️   {input_path} not found — multi-product pages skipped")

    # --- Sitemap ---
    generate_sitemap(blog_urls, geo_urls)

    # --- Submit ---
    if not args.no_submit:
        submit_sitemap()

    log.info("\n" + "=" * 55)
    log.info(f"🎉  DONE!")
    log.info(f"  📝  Blog posts:  {len(blog_urls)} URLs")
    log.info(f"  🗺️   Geo pages:   {len(geo_urls)} URLs")
    log.info(f"  📋  Sitemap:     {len(blog_urls) + len(geo_urls)} total URLs")
    log.info(f"  📂  Output:      {OUTPUT_DIR}")
    log.info(f"  🌐  Upload output/ folder to blog.adhiseal.com")
    log.info("=" * 55)


if __name__ == '__main__':
    main()
